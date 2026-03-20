"""
report_generation.py
--------------------
Airflow DAG: report_generation

Runs at 08:00 UTC on the 1st of each month.

Pipeline:
  1. fetch_report_data     — gather MERL summary data from PostgreSQL + ClickHouse
  2. generate_pdf          — create formatted monthly summary PDF with reportlab
  3. generate_excel        — create Excel data export with openpyxl
  4. save_reports          — persist report files to /opt/airflow/reports/
  5. email_reports         — send both files via SMTP to REPORT_RECIPIENTS
"""

from __future__ import annotations

import io
import logging
import os
import smtplib
import tempfile
from datetime import date, datetime, timedelta, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
POSTGRES_CONN_ID = "postgres_merl"
REPORTS_DIR = Path(os.getenv("REPORTS_DIR", "/opt/airflow/reports"))

CLICKHOUSE_HOST = os.getenv("CLICKHOUSE_HOST", "clickhouse")
CLICKHOUSE_PORT = int(os.getenv("CLICKHOUSE_PORT", "9000"))
CLICKHOUSE_DB = os.getenv("CLICKHOUSE_DB", "merl_analytics")
CLICKHOUSE_USER = os.getenv("CLICKHOUSE_USER", "merl_ch_user")
CLICKHOUSE_PASSWORD = os.getenv("CLICKHOUSE_PASSWORD", "")

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.example.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM = os.getenv("SMTP_FROM", "noreply@example.gov.vu")
REPORT_RECIPIENTS_RAW = os.getenv("REPORT_RECIPIENTS", "merl@docc.gov.vu")

PROJECT_NAME = "Vanuatu Loss and Damage Fund Development Project"
PROJECT_CODE = "VAN-CAP2-LNDD"

# ---------------------------------------------------------------------------
# Default args
# ---------------------------------------------------------------------------
default_args: dict[str, Any] = {
    "owner": "merl-team",
    "depends_on_past": False,
    "email_on_failure": False,
    "email_on_retry": False,
    "retries": 1,
    "retry_delay": timedelta(minutes=10),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_ch_client():
    from clickhouse_driver import Client  # type: ignore[import]

    return Client(
        host=CLICKHOUSE_HOST,
        port=CLICKHOUSE_PORT,
        database=CLICKHOUSE_DB,
        user=CLICKHOUSE_USER,
        password=CLICKHOUSE_PASSWORD,
        connect_timeout=10,
        send_receive_timeout=30,
    )


def _report_month(execution_date: datetime) -> tuple[date, date]:
    """Return (first_day, last_day) of the prior calendar month."""
    first_of_current = execution_date.replace(day=1).date()
    last_of_prior = first_of_current - timedelta(days=1)
    first_of_prior = last_of_prior.replace(day=1)
    return first_of_prior, last_of_prior


# ---------------------------------------------------------------------------
# Task callables
# ---------------------------------------------------------------------------

def fetch_report_data(**context: Any) -> None:
    """
    Gather MERL summary data from PostgreSQL and ClickHouse.
    Pushes a structured data dict via XCom.
    """
    execution_date: datetime = context["execution_date"]
    period_start, period_end = _report_month(execution_date)
    log.info("Generating report for period %s to %s", period_start, period_end)

    hook = PostgresHook(postgres_conn_id=POSTGRES_CONN_ID)

    # --- Project overview ---
    project_rows = hook.get_records(
        """
        SELECT
            COUNT(*)                                                    AS total_indicators,
            COUNT(*) FILTER (WHERE is_active)                           AS active_indicators,
            (SELECT COUNT(*) FROM merl.activities WHERE is_active)      AS total_activities,
            (SELECT COUNT(*) FROM merl.activities
             WHERE is_active AND status = 'completed')                  AS completed_activities,
            (SELECT COUNT(*) FROM merl.activities
             WHERE is_active AND status = 'in_progress')                AS inprogress_activities
        FROM merl.indicators;
        """
    )
    overview = dict(zip(
        ["total_indicators", "active_indicators", "total_activities",
         "completed_activities", "inprogress_activities"],
        project_rows[0],
    )) if project_rows else {}

    # --- KPI progress from ClickHouse ---
    kpi_rows: list[tuple] = []
    try:
        client = _get_ch_client()
        kpi_rows = client.execute(
            """
            SELECT
                indicator_code,
                indicator_name,
                domain,
                unit,
                current_value,
                target_value,
                progress_pct,
                trend,
                status
            FROM ch_merl.kpi_snapshots
            WHERE snapshot_date = (SELECT MAX(snapshot_date) FROM ch_merl.kpi_snapshots)
            ORDER BY domain, indicator_code;
            """
        )
    except Exception as exc:  # noqa: BLE001
        log.warning("ClickHouse KPI query failed: %s", exc)

    # --- Financial summary ---
    fin_rows = hook.get_records(
        """
        SELECT
            COALESCE(domain, 'Total')              AS domain,
            SUM(CASE WHEN transaction_type = 'allocation'   THEN amount ELSE 0 END) AS allocated,
            SUM(CASE WHEN transaction_type = 'disbursement' THEN amount ELSE 0 END) AS disbursed,
            SUM(CASE WHEN transaction_type = 'expenditure'  THEN amount ELSE 0 END) AS expended
        FROM   merl.financial_transactions
        WHERE  status != 'cancelled'
        GROUP  BY ROLLUP(domain)
        ORDER  BY domain NULLS LAST;
        """,
    )

    # --- Community engagement summary ---
    eng_rows = hook.get_records(
        """
        SELECT
            COUNT(*)                    AS total_events,
            SUM(total_participants)     AS total_participants,
            SUM(female_participants)    AS female_participants,
            SUM(youth_participants)     AS youth_participants,
            SUM(pwd_participants)       AS pwd_participants,
            COUNT(DISTINCT community_id) AS unique_communities
        FROM merl.community_engagements
        WHERE engagement_date BETWEEN %s AND %s;
        """,
        parameters=(period_start, period_end),
    )
    engagement = {}
    if eng_rows and eng_rows[0][0] is not None:
        row = eng_rows[0]
        total_p = int(row[1] or 0)
        female_p = int(row[2] or 0)
        youth_p = int(row[3] or 0)
        pwd_p = int(row[4] or 0)
        engagement = {
            "total_events": int(row[0] or 0),
            "total_participants": total_p,
            "female_participants": female_p,
            "youth_participants": youth_p,
            "pwd_participants": pwd_p,
            "unique_communities": int(row[5] or 0),
            "female_pct": round(female_p / total_p * 100, 1) if total_p > 0 else 0.0,
            "youth_pct": round(youth_p / total_p * 100, 1) if total_p > 0 else 0.0,
            "pwd_pct": round(pwd_p / total_p * 100, 1) if total_p > 0 else 0.0,
        }

    report_data = {
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "generated_at": datetime.now(tz=timezone.utc).isoformat(),
        "overview": overview,
        "kpi_rows": [list(r) for r in kpi_rows],
        "financial_rows": [list(r) for r in fin_rows],
        "engagement": engagement,
    }

    context["ti"].xcom_push(key="report_data", value=report_data)
    log.info("Report data fetched successfully.")


def generate_pdf(**context: Any) -> None:
    """
    Build a formatted PDF monthly report using reportlab.
    Pushes the PDF bytes (as list of ints) via XCom.
    """
    from reportlab.lib import colors  # type: ignore[import]
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, Spacer, Table, TableStyle, SimpleDocTemplate, HRFlowable,
    )

    ti = context["ti"]
    data: dict = ti.xcom_pull(task_ids="fetch_report_data", key="report_data") or {}

    period_start = data.get("period_start", "")
    period_end = data.get("period_end", "")
    overview = data.get("overview", {})
    kpi_rows = data.get("kpi_rows", [])
    fin_rows = data.get("financial_rows", [])
    engagement = data.get("engagement", {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    BRAND_BLUE = colors.HexColor("#1B4F72")
    BRAND_TEAL = colors.HexColor("#117A65")
    LIGHT_GREY = colors.HexColor("#F2F3F4")

    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        textColor=BRAND_BLUE, fontSize=18, spaceAfter=4,
    )
    heading1_style = ParagraphStyle(
        "H1", parent=styles["Heading1"],
        textColor=BRAND_BLUE, fontSize=13, spaceBefore=12, spaceAfter=4,
    )
    heading2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"],
        textColor=BRAND_TEAL, fontSize=11, spaceBefore=8, spaceAfter=2,
    )
    normal_style = styles["Normal"]
    small_style = ParagraphStyle("Small", parent=normal_style, fontSize=8)

    STATUS_COLOUR = {
        "on_track": colors.HexColor("#1E8449"),
        "at_risk": colors.HexColor("#D4AC0D"),
        "off_track": colors.HexColor("#CB4335"),
        "unknown": colors.grey,
    }

    elements = []

    # ---- Header ----
    elements.append(Paragraph(PROJECT_NAME, title_style))
    elements.append(Paragraph(f"Monthly MERL Report — {period_start} to {period_end}", heading2_style))
    elements.append(Paragraph(f"Generated: {data.get('generated_at', '')[:19]} UTC", small_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=BRAND_BLUE, spaceAfter=8))

    # ---- 1. Project Overview ----
    elements.append(Paragraph("1. Project Overview", heading1_style))
    overview_table_data = [
        ["Metric", "Value"],
        ["Total Indicators", str(overview.get("total_indicators", "—"))],
        ["Active Indicators", str(overview.get("active_indicators", "—"))],
        ["Total Activities", str(overview.get("total_activities", "—"))],
        ["Completed Activities", str(overview.get("completed_activities", "—"))],
        ["In-Progress Activities", str(overview.get("inprogress_activities", "—"))],
    ]
    overview_table = Table(overview_table_data, colWidths=[100 * mm, 60 * mm])
    overview_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(overview_table)

    # ---- 2. KPI Progress ----
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("2. KPI Progress", heading1_style))

    if kpi_rows:
        kpi_header = ["Code", "Indicator", "Domain", "Current", "Target", "Progress", "Trend", "Status"]
        kpi_table_data = [kpi_header]
        for r in kpi_rows:
            code, name, domain, unit, current, target, pct, trend, status = r
            name_short = str(name)[:50] + ("…" if len(str(name)) > 50 else "")
            progress_str = f"{pct:.1f}%" if pct is not None else "—"
            target_str = f"{target:,.1f} {unit}" if target is not None else "—"
            current_str = f"{current:,.1f} {unit}" if current is not None else "—"
            kpi_table_data.append([
                str(code), name_short, str(domain or ""),
                current_str, target_str, progress_str,
                str(trend or ""), str(status or ""),
            ])

        col_widths = [18 * mm, 55 * mm, 22 * mm, 22 * mm, 22 * mm, 18 * mm, 18 * mm, 18 * mm]
        kpi_table = Table(kpi_table_data, colWidths=col_widths, repeatRows=1)
        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("WORDWRAP", (1, 1), (1, -1), True),
        ])
        # Colour status column
        for i, row_data in enumerate(kpi_table_data[1:], start=1):
            status_val = row_data[-1].lower()
            cell_colour = STATUS_COLOUR.get(status_val, colors.white)
            ts.add("BACKGROUND", (7, i), (7, i), cell_colour)
            if status_val != "unknown":
                ts.add("TEXTCOLOR", (7, i), (7, i), colors.white)

        kpi_table.setStyle(ts)
        elements.append(kpi_table)
    else:
        elements.append(Paragraph("No KPI snapshot data available for this period.", normal_style))

    # ---- 3. Financial Summary ----
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("3. Financial Summary (USD)", heading1_style))

    if fin_rows:
        fin_header = ["Domain", "Allocated", "Disbursed", "Expended", "Disb. Rate", "Exp. Rate"]
        fin_table_data = [fin_header]
        for r in fin_rows:
            domain, allocated, disbursed, expended = r
            allocated = float(allocated or 0)
            disbursed = float(disbursed or 0)
            expended = float(expended or 0)
            disb_rate = f"{disbursed / allocated * 100:.1f}%" if allocated > 0 else "—"
            exp_rate = f"{expended / allocated * 100:.1f}%" if allocated > 0 else "—"
            fin_table_data.append([
                str(domain or "Total"),
                f"${allocated:,.0f}",
                f"${disbursed:,.0f}",
                f"${expended:,.0f}",
                disb_rate,
                exp_rate,
            ])
        fin_table = Table(fin_table_data, colWidths=[40 * mm, 28 * mm, 28 * mm, 28 * mm, 22 * mm, 22 * mm])
        fin_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            # Bold the totals row (last row if it has no domain)
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), LIGHT_GREY),
        ]))
        elements.append(fin_table)
    else:
        elements.append(Paragraph("No financial data available.", normal_style))

    # ---- 4. Community Engagement & GEDSI ----
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("4. Community Engagement & GEDSI", heading1_style))
    elements.append(Paragraph(
        f"Reporting period: {period_start} to {period_end}", small_style,
    ))
    elements.append(Spacer(1, 2 * mm))

    if engagement:
        gedsi_data = [
            ["Metric", "Count", "% of Total Participants"],
            ["Total Engagement Events", str(engagement.get("total_events", 0)), "—"],
            ["Total Participants", str(f"{engagement.get('total_participants', 0):,}"), "100%"],
            ["Female Participants", str(f"{engagement.get('female_participants', 0):,}"),
             f"{engagement.get('female_pct', 0):.1f}%"],
            ["Youth Participants", str(f"{engagement.get('youth_participants', 0):,}"),
             f"{engagement.get('youth_pct', 0):.1f}%"],
            ["Persons with Disabilities", str(f"{engagement.get('pwd_participants', 0):,}"),
             f"{engagement.get('pwd_pct', 0):.1f}%"],
            ["Unique Communities Reached", str(engagement.get("unique_communities", 0)), "—"],
        ]
        gedsi_table = Table(gedsi_data, colWidths=[80 * mm, 40 * mm, 55 * mm])
        gedsi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_TEAL),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(gedsi_table)
    else:
        elements.append(Paragraph("No community engagement data recorded for this period.", normal_style))

    # ---- Footer ----
    elements.append(Spacer(1, 10 * mm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    elements.append(Paragraph(
        f"Confidential — {PROJECT_CODE} | Department of Climate Change | Vanuatu",
        ParagraphStyle("Footer", parent=small_style, textColor=colors.grey, alignment=1),
    ))

    doc.build(elements)
    pdf_bytes = buf.getvalue()
    context["ti"].xcom_push(key="pdf_bytes", value=list(pdf_bytes))
    log.info("PDF generated: %d bytes.", len(pdf_bytes))


def generate_excel(**context: Any) -> None:
    """
    Create an Excel workbook with multiple sheets for the monthly export.
    Pushes Excel bytes (as list of ints) via XCom.
    """
    import openpyxl  # type: ignore[import]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ti = context["ti"]
    data: dict = ti.xcom_pull(task_ids="fetch_report_data", key="report_data") or {}

    period_start = data.get("period_start", "")
    period_end = data.get("period_end", "")
    kpi_rows = data.get("kpi_rows", [])
    fin_rows = data.get("financial_rows", [])
    engagement = data.get("engagement", {})
    overview = data.get("overview", {})

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    ALT_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    BOLD_FONT = Font(bold=True, size=10)
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def apply_header_row(ws, row_idx: int, values: list) -> None:
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def auto_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # ---- Sheet 1: Summary ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = PROJECT_NAME
    ws_summary["A1"].font = Font(bold=True, size=14, color="1B4F72")
    ws_summary["A2"] = f"Monthly MERL Report — {period_start} to {period_end}"
    ws_summary["A2"].font = Font(italic=True, size=11)
    ws_summary["A3"] = f"Generated: {data.get('generated_at', '')[:19]} UTC"

    row = 5
    apply_header_row(ws_summary, row, ["Metric", "Value"])
    metrics = [
        ("Period", f"{period_start} to {period_end}"),
        ("Total Indicators", overview.get("total_indicators", 0)),
        ("Active Indicators", overview.get("active_indicators", 0)),
        ("Total Activities", overview.get("total_activities", 0)),
        ("Completed Activities", overview.get("completed_activities", 0)),
        ("In-Progress Activities", overview.get("inprogress_activities", 0)),
    ]
    for i, (label, value) in enumerate(metrics, start=row + 1):
        ws_summary.cell(row=i, column=1, value=label).border = THIN_BORDER
        ws_summary.cell(row=i, column=2, value=value).border = THIN_BORDER
        if i % 2 == 0:
            for c in range(1, 3):
                ws_summary.cell(row=i, column=c).fill = ALT_FILL
    auto_width(ws_summary)

    # ---- Sheet 2: KPI Progress ----
    ws_kpi = wb.create_sheet("KPI Progress")
    kpi_headers = ["Code", "Indicator Name", "Domain", "Unit", "Current Value",
                   "Target Value", "Progress %", "Trend", "Status", "Last Reported"]
    apply_header_row(ws_kpi, 1, kpi_headers)
    ws_kpi.freeze_panes = "A2"
    for i, r in enumerate(kpi_rows, start=2):
        code, name, domain, unit, current, target, pct, trend, status = r
        values = [code, name, domain, unit, current, target,
                  round(pct, 2) if pct is not None else None, trend, status, ""]
        for col_idx, val in enumerate(values, start=1):
            cell = ws_kpi.cell(row=i, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
    auto_width(ws_kpi)

    # ---- Sheet 3: Financial Summary ----
    ws_fin = wb.create_sheet("Financial Summary")
    fin_headers = ["Domain", "Allocated (USD)", "Disbursed (USD)", "Expended (USD)",
                   "Disbursement Rate %", "Expenditure Rate %"]
    apply_header_row(ws_fin, 1, fin_headers)
    ws_fin.freeze_panes = "A2"
    for i, r in enumerate(fin_rows, start=2):
        domain, allocated, disbursed, expended = r
        allocated = float(allocated or 0)
        disbursed = float(disbursed or 0)
        expended = float(expended or 0)
        disb_rate = round(disbursed / allocated * 100, 2) if allocated > 0 else None
        exp_rate = round(expended / allocated * 100, 2) if allocated > 0 else None
        row_vals = [domain, allocated, disbursed, expended, disb_rate, exp_rate]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_fin.cell(row=i, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
    auto_width(ws_fin)

    # ---- Sheet 4: GEDSI Engagement ----
    ws_eng = wb.create_sheet("GEDSI Engagement")
    eng_headers = ["Metric", "Value", "% of Participants"]
    apply_header_row(ws_eng, 1, eng_headers)
    total_p = engagement.get("total_participants", 0)
    eng_metrics = [
        ("Total Events", engagement.get("total_events", 0), None),
        ("Total Participants", total_p, "100%"),
        ("Female Participants", engagement.get("female_participants", 0),
         f"{engagement.get('female_pct', 0):.1f}%"),
        ("Youth Participants", engagement.get("youth_participants", 0),
         f"{engagement.get('youth_pct', 0):.1f}%"),
        ("Persons with Disabilities", engagement.get("pwd_participants", 0),
         f"{engagement.get('pwd_pct', 0):.1f}%"),
        ("Unique Communities", engagement.get("unique_communities", 0), None),
    ]
    for i, (label, value, pct) in enumerate(eng_metrics, start=2):
        ws_eng.cell(row=i, column=1, value=label).border = THIN_BORDER
        ws_eng.cell(row=i, column=2, value=value).border = THIN_BORDER
        ws_eng.cell(row=i, column=3, value=pct or "—").border = THIN_BORDER
        if i % 2 == 0:
            for c in range(1, 4):
                ws_eng.cell(row=i, column=c).fill = ALT_FILL
    auto_width(ws_eng)

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    context["ti"].xcom_push(key="excel_bytes", value=list(excel_bytes))
    log.info("Excel workbook generated: %d bytes.", len(excel_bytes))


def save_reports(**context: Any) -> None:
    """
    Write PDF and Excel files to REPORTS_DIR and push file paths via XCom.
    """
    ti = context["ti"]
    data: dict = ti.xcom_pull(task_ids="fetch_report_data", key="report_data") or {}
    pdf_bytes_list: list[int] = ti.xcom_pull(task_ids="generate_pdf", key="pdf_bytes") or []
    excel_bytes_list: list[int] = ti.xcom_pull(task_ids="generate_excel", key="excel_bytes") or []

    period_str = data.get("period_start", date.today().strftime("%Y-%m"))[:7].replace("-", "_")
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    pdf_path = REPORTS_DIR / f"merl_monthly_report_{period_str}.pdf"
    excel_path = REPORTS_DIR / f"merl_monthly_data_{period_str}.xlsx"

    if pdf_bytes_list:
        pdf_path.write_bytes(bytes(pdf_bytes_list))
        log.info("PDF saved to %s (%d bytes)", pdf_path, len(pdf_bytes_list))

    if excel_bytes_list:
        excel_path.write_bytes(bytes(excel_bytes_list))
        log.info("Excel saved to %s (%d bytes)", excel_path, len(excel_bytes_list))

    ti.xcom_push(key="pdf_path", value=str(pdf_path))
    ti.xcom_push(key="excel_path", value=str(excel_path))


def email_reports(**context: Any) -> None:
    """
    Send monthly report files via SMTP to REPORT_RECIPIENTS.
    Skips gracefully if SMTP_HOST is not configured.
    """
    ti = context["ti"]
    data: dict = ti.xcom_pull(task_ids="fetch_report_data", key="report_data") or {}
    pdf_path_str: str = ti.xcom_pull(task_ids="save_reports", key="pdf_path") or ""
    excel_path_str: str = ti.xcom_pull(task_ids="save_reports", key="excel_path") or ""

    recipients = [r.strip() for r in REPORT_RECIPIENTS_RAW.split(",") if r.strip()]
    if not recipients:
        log.warning("No REPORT_RECIPIENTS configured; skipping email.")
        return

    if SMTP_HOST in ("", "smtp.example.com"):
        log.warning("SMTP_HOST not configured; skipping email send.")
        return

    period_start = data.get("period_start", "")
    subject = f"[MERL Dashboard] Monthly Report — {period_start[:7]}"
    body = (
        f"Please find attached the MERL Monthly Report for {period_start[:7]}.\n\n"
        f"Project: {PROJECT_NAME}\n"
        f"Period: {period_start} to {data.get('period_end', '')}\n"
        f"Generated: {data.get('generated_at', '')[:19]} UTC\n\n"
        "This is an automated message from the MERL Dashboard system.\n"
        "Do not reply to this email."
    )

    msg = MIMEMultipart()
    msg["From"] = SMTP_FROM
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    for file_path_str in [pdf_path_str, excel_path_str]:
        if not file_path_str:
            continue
        file_path = Path(file_path_str)
        if not file_path.exists():
            log.warning("Attachment not found: %s", file_path)
            continue
        with file_path.open("rb") as fh:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(fh.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{file_path.name}"')
        msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
            server.ehlo()
            server.starttls()
            if SMTP_USER and SMTP_PASSWORD:
                server.login(SMTP_USER, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM, recipients, msg.as_string())
        log.info("Monthly report emailed to: %s", ", ".join(recipients))
    except Exception as exc:  # noqa: BLE001
        log.error("Failed to send email: %s", exc)
        raise


# ---------------------------------------------------------------------------
# DAG definition
# ---------------------------------------------------------------------------

with DAG(
    dag_id="report_generation",
    description="Monthly MERL report generation: PDF + Excel, saved to disk and emailed",
    default_args=default_args,
    schedule_interval="0 8 1 * *",
    start_date=datetime(2024, 1, 1, tzinfo=timezone.utc),
    catchup=False,
    max_active_runs=1,
    tags=["merl", "reports", "monthly"],
) as dag:

    t_fetch = PythonOperator(
        task_id="fetch_report_data",
        python_callable=fetch_report_data,
    )

    t_pdf = PythonOperator(
        task_id="generate_pdf",
        python_callable=generate_pdf,
    )

    t_excel = PythonOperator(
        task_id="generate_excel",
        python_callable=generate_excel,
    )

    t_save = PythonOperator(
        task_id="save_reports",
        python_callable=save_reports,
    )

    t_email = PythonOperator(
        task_id="email_reports",
        python_callable=email_reports,
    )

    t_fetch >> [t_pdf, t_excel] >> t_save >> t_email
