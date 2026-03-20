"""
Export service: PDF and Excel report generation, and email delivery.

Functions:
  - generate_pdf_report(summary_data)  → bytes
  - generate_excel_report(data_dict)   → bytes
  - email_report(recipients, pdf_bytes) → None
"""

from __future__ import annotations

import io
import logging
import smtplib
from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.config import settings

logger = logging.getLogger(__name__)

# ── Color constants ───────────────────────────────────────────────────────────

_BRAND_BLUE = colors.HexColor("#1a5276")
_BRAND_GREEN = colors.HexColor("#1e8449")
_LIGHT_GREY = colors.HexColor("#f2f3f4")
_DARK_GREY = colors.HexColor("#566573")

# ── PDF generation ────────────────────────────────────────────────────────────


def generate_pdf_report(summary_data: Dict[str, Any]) -> bytes:
    """
    Generate a structured PDF MERL summary report using ReportLab.

    Args:
        summary_data: Dict produced by reports.py _build_merl_summary().

    Returns:
        Raw PDF bytes ready to stream to the client or attach to an email.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2 * cm,
        title="MERL Dashboard Summary Report",
        author="VCAP2 Project",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        fontSize=18,
        textColor=_BRAND_BLUE,
        spaceAfter=6,
    )
    heading1_style = ParagraphStyle(
        "Heading1Style",
        parent=styles["Heading1"],
        fontSize=13,
        textColor=_BRAND_BLUE,
        spaceBefore=12,
        spaceAfter=4,
    )
    heading2_style = ParagraphStyle(
        "Heading2Style",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=_BRAND_GREEN,
        spaceBefore=8,
        spaceAfter=3,
    )
    body_style = styles["BodyText"]
    body_style.fontSize = 9

    story = []

    # ── Cover / Header ────────────────────────────────────────────────────────
    story.append(
        Paragraph(
            "Vanuatu Loss & Damage Fund Development Project",
            title_style,
        )
    )
    story.append(
        Paragraph("MERL Dashboard – Summary Report", heading1_style)
    )
    story.append(
        Paragraph(f"Generated: {date.today().strftime('%d %B %Y')}", body_style)
    )
    story.append(HRFlowable(width="100%", thickness=1, color=_BRAND_BLUE))
    story.append(Spacer(1, 0.5 * cm))

    # ── Indicators section ────────────────────────────────────────────────────
    ind_data = summary_data.get("indicators", {})
    story.append(Paragraph("1. Indicator Status", heading1_style))
    story.append(
        Paragraph(
            f"Total active indicators: <b>{ind_data.get('total_active', 0)}</b>",
            body_style,
        )
    )
    by_status = ind_data.get("by_status", {})
    if by_status:
        table_data = [["Status", "Count"]] + [
            [k.replace("_", " ").title(), str(v)] for k, v in by_status.items()
        ]
        story.append(_build_table(table_data))
    story.append(Spacer(1, 0.3 * cm))

    # ── Activities section ────────────────────────────────────────────────────
    act_data = summary_data.get("activities", {})
    story.append(Paragraph("2. Activity Progress", heading1_style))
    story.append(
        Paragraph(
            f"Total active activities: <b>{act_data.get('total_active', 0)}</b> &nbsp; "
            f"Completion rate: <b>{act_data.get('completion_rate_pct', 0):.1f}%</b>",
            body_style,
        )
    )
    act_by_status = act_data.get("by_status", {})
    if act_by_status:
        table_data = [["Status", "Count"]] + [
            [k.replace("_", " ").title(), str(v)]
            for k, v in act_by_status.items()
        ]
        story.append(_build_table(table_data))
    story.append(Spacer(1, 0.3 * cm))

    # ── Financials section ────────────────────────────────────────────────────
    fin_data = summary_data.get("financials", {})
    story.append(Paragraph("3. Financial Summary", heading1_style))
    disbursed = fin_data.get("total_disbursed", 0.0)
    spent = fin_data.get("total_spent", 0.0)
    fin_table = [
        ["Metric", "Amount (VUV)"],
        ["Total Disbursed", f"{disbursed:,.2f}"],
        ["Total Spent", f"{spent:,.2f}"],
        [
            "Absorption Rate",
            f"{round(spent / disbursed * 100, 1) if disbursed else 0:.1f}%",
        ],
    ]
    story.append(_build_table(fin_table))
    story.append(Spacer(1, 0.3 * cm))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=_DARK_GREY))
    story.append(Spacer(1, 0.2 * cm))
    story.append(
        Paragraph(
            "CONFIDENTIAL – For internal project use only.",
            ParagraphStyle(
                "Footer",
                parent=body_style,
                fontSize=7,
                textColor=_DARK_GREY,
            ),
        )
    )

    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def _build_table(data: List[List[str]]) -> Table:
    """Create a styled ReportLab table from a 2D list."""
    tbl = Table(data, hAlign="LEFT")
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _BRAND_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _LIGHT_GREY]),
                ("GRID", (0, 0), (-1, -1), 0.4, _DARK_GREY),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return tbl


# ── Excel generation ──────────────────────────────────────────────────────────

_XL_HEADER_FILL = PatternFill("solid", fgColor="1A5276")
_XL_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_XL_ALT_FILL = PatternFill("solid", fgColor="EAF2F8")


def generate_excel_report(data_dict: Dict[str, Any]) -> bytes:
    """
    Generate an Excel MERL report workbook with multiple sheets.

    Args:
        data_dict: Dict with keys 'summary', 'indicators', 'activities',
                   'transactions'.  Each value is a list of flat dicts.

    Returns:
        Raw .xlsx bytes.
    """
    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary = data_dict.get("summary", {})
    ws_summary.append(["MERL Dashboard – Summary Report"])
    ws_summary.append([f"Generated: {date.today().strftime('%d %B %Y')}"])
    ws_summary.append([])

    ws_summary.append(["Section", "Key", "Value"])
    _style_header_row(ws_summary, 4)

    for section, values in summary.items():
        if isinstance(values, dict):
            for k, v in values.items():
                ws_summary.append([section, k, str(v)])
        else:
            ws_summary.append([section, "", str(values)])

    _autofit_columns(ws_summary)

    # ── Indicators sheet ──────────────────────────────────────────────────────
    indicators = data_dict.get("indicators", [])
    if indicators:
        ws_ind = wb.create_sheet("Indicators")
        headers = list(indicators[0].keys())
        ws_ind.append(headers)
        _style_header_row(ws_ind, 1)
        for i, row in enumerate(indicators):
            ws_ind.append([row.get(h, "") for h in headers])
            if i % 2 == 0:
                for cell in ws_ind[ws_ind.max_row]:
                    cell.fill = _XL_ALT_FILL
        _autofit_columns(ws_ind)

    # ── Activities sheet ──────────────────────────────────────────────────────
    activities = data_dict.get("activities", [])
    if activities:
        ws_act = wb.create_sheet("Activities")
        headers = list(activities[0].keys())
        ws_act.append(headers)
        _style_header_row(ws_act, 1)
        for i, row in enumerate(activities):
            ws_act.append([row.get(h, "") for h in headers])
            if i % 2 == 0:
                for cell in ws_act[ws_act.max_row]:
                    cell.fill = _XL_ALT_FILL
        _autofit_columns(ws_act)

    # ── Transactions sheet ────────────────────────────────────────────────────
    transactions = data_dict.get("transactions", [])
    if transactions:
        ws_tx = wb.create_sheet("Transactions")
        headers = list(transactions[0].keys())
        ws_tx.append(headers)
        _style_header_row(ws_tx, 1)
        for i, row in enumerate(transactions):
            ws_tx.append([row.get(h, "") for h in headers])
            if i % 2 == 0:
                for cell in ws_tx[ws_tx.max_row]:
                    cell.fill = _XL_ALT_FILL
        _autofit_columns(ws_tx)

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()
    buffer.close()
    return xlsx_bytes


def _style_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, row: int) -> None:
    """Apply header styling to a given row in an openpyxl worksheet."""
    for cell in ws[row]:
        cell.fill = _XL_HEADER_FILL
        cell.font = _XL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Set a reasonable column width based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


# ── Email delivery ────────────────────────────────────────────────────────────


def email_report(
    recipients: List[str],
    pdf_bytes: bytes,
    subject: str = "MERL Dashboard – Automated Summary Report",
    body_text: str | None = None,
) -> None:
    """
    Send a PDF report as an email attachment via SMTP.

    Args:
        recipients: List of recipient email addresses.
        pdf_bytes:  Raw PDF bytes to attach.
        subject:    Email subject line.
        body_text:  Optional plain-text body.  A default is used if None.

    Raises:
        smtplib.SMTPException: On SMTP protocol errors.
        OSError: On network-level connection errors.
    """
    if not recipients:
        logger.warning("email_report called with empty recipient list – skipping.")
        return

    if body_text is None:
        body_text = (
            f"Please find attached the automated MERL Summary Report "
            f"generated on {date.today().strftime('%d %B %Y')}.\n\n"
            "This message was sent automatically by the MERL Dashboard.\n"
            "Do not reply to this email."
        )

    msg = MIMEMultipart()
    msg["From"] = settings.SMTP_FROM
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(body_text, "plain", "utf-8"))

    attachment = MIMEApplication(pdf_bytes, _subtype="pdf")
    attachment.add_header(
        "Content-Disposition",
        "attachment",
        filename="merl_summary_report.pdf",
    )
    msg.attach(attachment)

    try:
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT, timeout=30) as server:
            server.ehlo()
            if settings.SMTP_PORT in (587, 2587):
                server.starttls()
                server.ehlo()
            if settings.SMTP_USER and settings.SMTP_PASSWORD:
                server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            server.sendmail(settings.SMTP_FROM, recipients, msg.as_string())
        logger.info(
            "MERL report email sent to %d recipient(s): %s",
            len(recipients),
            recipients,
        )
    except smtplib.SMTPException as exc:
        logger.error("Failed to send MERL report email: %s", exc)
        raise
